"""Exports (§30) — the writers, and the two endpoints that use them.

The interesting assertions are not "it produced a file". They are the four
things an export written by hand gets wrong: it exports the page instead of
the question, it holds the whole result in memory, it corrupts every accented
name in Excel, and it hands Excel a formula somebody typed into a text box.
"""

from __future__ import annotations

import csv
import io
import json
from uuid import uuid4

import pytest

from src.config import Config
from src.core import export

PREFIX = Config.API_PREFIX

COLUMNS = [export.Column("reference", "Reference"), export.Column("title")]


def _claims(username: str, role: str) -> dict:
    return {
        "sub": "", "email": f"{username}@nucleus.example",
        "preferred_username": username, "name": username.title(),
        "sid": f"export-{username}", "realm_access": {"roles": [role]},
    }


def _authenticate(monkeypatch, username: str = "admin", role: str = "administrator"):
    monkeypatch.setattr("src.core.auth.verify_token", lambda _token: _claims(username, role))
    return {"Authorization": f"Bearer export-{username}"}


# ── the writers ──────────────────────────────────────────────────────────


def test_csv_starts_with_a_bom_so_excel_reads_utf8():
    # Without it, Excel on Windows reads the file as the local codepage and
    # every accented name in it comes out mangled.
    body = "".join(export.csv_lines([{"reference": "TSK-1", "title": "Café résumé"}], COLUMNS))

    assert body.startswith("﻿")
    assert "Café résumé" in body


def test_csv_quotes_anything_a_spreadsheet_would_execute():
    rows = [
        {"reference": "=cmd|'/c calc'!A1", "title": "+1234"},
        {"reference": "@SUM(1)", "title": "-alarming"},
        {"reference": "ordinary", "title": "also ordinary"},
    ]

    body = "".join(export.csv_lines(rows, COLUMNS))
    parsed = list(csv.reader(io.StringIO(body.lstrip("﻿"))))

    assert parsed[1] == ["'=cmd|'/c calc'!A1", "'+1234"]
    assert parsed[2] == ["'@SUM(1)", "'-alarming"]
    # An ordinary value is left exactly as it was: a quote in front of every
    # cell is its own kind of broken export.
    assert parsed[3] == ["ordinary", "also ordinary"]


def test_csv_leaves_a_negative_number_a_number():
    # `cell` keeps numbers numeric, so `defuse` never sees them and a numeric
    # column does not silently become text.
    body = "".join(export.csv_lines([{"reference": -3, "title": 4.5}], COLUMNS))
    parsed = list(csv.reader(io.StringIO(body.lstrip("﻿"))))

    assert parsed[1] == ["-3", "4.5"]


def test_json_is_a_real_array_and_keeps_null_as_null():
    rows = [{"reference": "TSK-1", "title": None}, {"reference": "TSK-2", "title": "second"}]

    parsed = json.loads("".join(export.json_lines(rows, COLUMNS)))

    assert parsed == [
        {"reference": "TSK-1", "title": None},
        {"reference": "TSK-2", "title": "second"},
    ]


def test_json_of_nothing_is_an_empty_array_not_an_empty_file():
    assert json.loads("".join(export.json_lines([], COLUMNS))) == []


def test_xlsx_is_a_workbook_with_the_declared_headings():
    from openpyxl import load_workbook

    blob = export.xlsx_bytes([{"reference": "TSK-1", "title": "First"}], COLUMNS)
    sheet = load_workbook(io.BytesIO(blob)).active

    assert [cell.value for cell in sheet[1]] == ["Reference", "Title"]
    assert [cell.value for cell in sheet[2]] == ["TSK-1", "First"]


def test_a_filename_says_what_it_holds_and_sorts_by_when():
    name = export.filename("audit log/2026", "csv")

    assert name.startswith("audit-log-2026-")
    assert name.endswith(".csv")


def test_an_unknown_format_is_a_client_error():
    from src.core.errors import ValidationError

    with pytest.raises(ValidationError):
        export.parse_format("pdf")
    assert export.parse_format(None) == "csv"
    assert export.parse_format("XLSX") == "xlsx"


def test_xlsx_is_capped_lower_than_the_streaming_formats():
    # It has to be complete before the first byte goes out, so the ceiling is
    # what one worker should hold rather than what a stream can carry.
    assert export.limit_for("xlsx") < export.limit_for("csv")
    assert export.limit_for("json") == export.MAX_ROWS


# ── the endpoints ────────────────────────────────────────────────────────


def test_export_endpoints_require_a_bearer_token(client):
    assert client.get(f"{PREFIX}/admin/audit/export").status_code == 401
    assert client.post(f"{PREFIX}/api/explorer/export", json={}).status_code == 401


@pytest.mark.database
def test_taking_a_copy_away_is_a_separate_permission_from_reading(client, monkeypatch):
    # VIEWER may read records and may not export them. Reading the ledger and
    # walking out of the building with it are different privileges.
    headers = _authenticate(monkeypatch, "user", "viewer")

    refused = client.post(
        f"{PREFIX}/api/explorer/export", headers=headers,
        json={"resource_type": "task", "format": "csv"},
    )

    assert refused.status_code == 403
    assert "records.export" in refused.get_json()["details"]["missing"]


@pytest.mark.database
def test_the_explorer_exports_the_question_rather_than_the_page(client, monkeypatch):
    headers = _authenticate(monkeypatch)
    question = {
        "resource_type": "task",
        "filters": {"status": "DONE"},
        "columns": ["reference", "title", "status"],
        # A page size the file must ignore: this is the whole point.
        "page_size": 5,
    }

    counted = client.post(
        f"{PREFIX}/api/explorer/query", headers=headers, json=question
    ).get_json()

    response = client.post(
        f"{PREFIX}/api/explorer/export", headers=headers, json={**question, "format": "csv"}
    )

    assert response.status_code == 200
    assert response.mimetype == "text/csv"
    assert "attachment; filename=" in response.headers["Content-Disposition"]

    body = response.get_data(as_text=True).lstrip("﻿")
    rows = list(csv.reader(io.StringIO(body)))
    assert rows[0] == ["Reference", "Title", "Status"]
    # Every matching row, not the five on the page.
    assert len(rows) - 1 == counted["total"]
    assert counted["total"] > 5
    assert {row[2] for row in rows[1:]} == {"DONE"}


@pytest.mark.database
def test_the_explorer_export_honours_the_requested_format(client, monkeypatch):
    from openpyxl import load_workbook

    headers = _authenticate(monkeypatch)
    question = {
        "resource_type": "task",
        "filters": {"status": "DONE"},
        "columns": ["reference", "title"],
    }

    as_json = client.post(
        f"{PREFIX}/api/explorer/export", headers=headers, json={**question, "format": "json"}
    )
    as_xlsx = client.post(
        f"{PREFIX}/api/explorer/export", headers=headers, json={**question, "format": "xlsx"}
    )

    parsed = json.loads(as_json.get_data(as_text=True))
    assert isinstance(parsed, list) and parsed
    assert set(parsed[0]) == {"reference", "title"}

    assert as_xlsx.mimetype == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    sheet = load_workbook(io.BytesIO(as_xlsx.get_data())).active
    assert [cell.value for cell in sheet[1]] == ["Reference", "Title"]


@pytest.mark.database
def test_an_unknown_export_column_is_refused_before_anything_is_written(client, monkeypatch):
    response = client.post(
        f"{PREFIX}/api/explorer/export", headers=_authenticate(monkeypatch),
        json={"resource_type": "task", "columns": ["reference", "salary"]},
    )

    assert response.status_code == 400
    assert response.get_json()["details"]["columns"] == ["salary"]


@pytest.mark.database
def test_the_audit_ledger_exports_under_the_filters_on_screen(client, monkeypatch):
    headers = _authenticate(monkeypatch)
    query = "action=UPDATE&result=SUCCESS"

    listed = client.get(f"{PREFIX}/admin/audit?{query}&page_size=1", headers=headers).get_json()
    response = client.get(f"{PREFIX}/admin/audit/export?{query}&format=csv", headers=headers)

    assert response.status_code == 200
    assert "audit-log-" in response.headers["Content-Disposition"]

    rows = list(csv.reader(io.StringIO(response.get_data(as_text=True).lstrip("﻿"))))
    assert len(rows) - 1 == listed["total"]
    assert "Correlation ID" in rows[0]
    # The columns a table has no room for but an investigation needs.
    assert "Impersonator" in rows[0]


@pytest.mark.database
def test_the_audit_export_needs_both_reading_and_exporting(client, monkeypatch):
    # ANALYST reads the ledger and may export; OPERATOR may export records but
    # cannot read the ledger at all.
    analyst = client.get(
        f"{PREFIX}/admin/audit/export?format=csv&page_size=1",
        headers=_authenticate(monkeypatch, "analyst", "analyst"),
    )
    operator = client.get(
        f"{PREFIX}/admin/audit/export?format=csv",
        headers=_authenticate(monkeypatch, "operator", "operator"),
    )

    assert analyst.status_code == 200
    assert operator.status_code == 403
    assert operator.get_json()["details"]["missing"] == ["audit.view"]


@pytest.mark.database
def test_an_export_of_nothing_is_a_file_with_only_headings(client, monkeypatch):
    response = client.get(
        f"{PREFIX}/admin/audit/export?format=csv&correlation_id={uuid4().hex}",
        headers=_authenticate(monkeypatch),
    )

    rows = list(csv.reader(io.StringIO(response.get_data(as_text=True).lstrip("﻿"))))
    assert response.status_code == 200
    assert len(rows) == 1
    assert rows[0][0] == "When"
